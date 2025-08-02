#!/usr/bin/env python3
"""
Excel Workbook Branding Applicator

Applies advanced branding and styling to Excel workbooks with color schemes,
fonts, conditional formatting, and professional styling.

Features:
- Color scheme application to cells, charts, and backgrounds
- Font styling for different worksheet elements
- Conditional formatting for compliance status visualization
- Professional table styling with alternating row colors
- Chart color coordination with brand colors
"""

import logging
from typing import Dict, List, Any, Optional, Tuple

from .branding_system import (
    BrandingApplicator, AdvancedBrandingConfig, ColorUtilities, 
    ConditionalFormattingTheme
)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.styles.colors import Color
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
    from openpyxl.chart import BarChart, PieChart, LineChart
    from openpyxl.chart.series import DataPoint
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    # Create placeholders for type hints
    openpyxl = Any
    Font = Any
    PatternFill = Any
    Border = Any
    Side = Any
    Alignment = Any


class ExcelBrandingApplicator(BrandingApplicator):
    """Applies branding to Excel workbooks."""
    
    def __init__(self, branding_config: AdvancedBrandingConfig):
        super().__init__(branding_config)
        
        if not OPENPYXL_AVAILABLE:
            self.logger.error("openpyxl library not available")
            raise ImportError("openpyxl library required for Excel branding")
        
        # Pre-calculate style objects
        self._initialize_styles()
    
    def _initialize_styles(self):
        """Initialize reusable style objects."""
        self.styles = {
            # Fonts
            'title_font': Font(
                name=self.branding.fonts.primary_font,
                size=self.branding.fonts.title_size,
                bold=True,
                color=self._hex_to_openpyxl_color(self.branding.colors.primary)
            ),
            'heading_font': Font(
                name=self.branding.fonts.primary_font,
                size=self.branding.fonts.heading1_size,
                bold=True,
                color=self._hex_to_openpyxl_color(self.branding.colors.primary)
            ),
            'header_font': Font(
                name=self.branding.fonts.primary_font,
                size=self.branding.fonts.table_size,
                bold=True,
                color=self._hex_to_openpyxl_color(self.branding.colors.header_text)
            ),
            'body_font': Font(
                name=self.branding.fonts.primary_font,
                size=self.branding.fonts.body_size
            ),
            'table_font': Font(
                name=self.branding.fonts.primary_font,
                size=self.branding.fonts.table_size
            ),
            
            # Fills
            'header_fill': PatternFill(
                start_color=self._hex_to_openpyxl_color(self.branding.colors.header_bg),
                end_color=self._hex_to_openpyxl_color(self.branding.colors.header_bg),
                fill_type='solid'
            ),
            'alt_row_fill': PatternFill(
                start_color=self._hex_to_openpyxl_color(self.branding.colors.alt_row),
                end_color=self._hex_to_openpyxl_color(self.branding.colors.alt_row),
                fill_type='solid'
            ),
            'compliant_fill': PatternFill(
                start_color=self._hex_to_openpyxl_color(ColorUtilities.lighten_color(self.branding.colors.compliant, 0.8)),
                end_color=self._hex_to_openpyxl_color(ColorUtilities.lighten_color(self.branding.colors.compliant, 0.8)),
                fill_type='solid'
            ),
            'non_compliant_fill': PatternFill(
                start_color=self._hex_to_openpyxl_color(ColorUtilities.lighten_color(self.branding.colors.non_compliant, 0.8)),
                end_color=self._hex_to_openpyxl_color(ColorUtilities.lighten_color(self.branding.colors.non_compliant, 0.8)),
                fill_type='solid'
            ),
            
            # Borders
            'thin_border': Border(
                left=Side(style='thin', color=self._hex_to_openpyxl_color(self.branding.colors.border)),
                right=Side(style='thin', color=self._hex_to_openpyxl_color(self.branding.colors.border)),
                top=Side(style='thin', color=self._hex_to_openpyxl_color(self.branding.colors.border)),
                bottom=Side(style='thin', color=self._hex_to_openpyxl_color(self.branding.colors.border))
            ),
            'thick_border': Border(
                left=Side(style='thick', color=self._hex_to_openpyxl_color(self.branding.colors.primary)),
                right=Side(style='thick', color=self._hex_to_openpyxl_color(self.branding.colors.primary)),
                top=Side(style='thick', color=self._hex_to_openpyxl_color(self.branding.colors.primary)),
                bottom=Side(style='thick', color=self._hex_to_openpyxl_color(self.branding.colors.primary))
            ),
            
            # Alignments
            'center_alignment': Alignment(horizontal='center', vertical='center'),
            'left_alignment': Alignment(horizontal='left', vertical='center'),
            'right_alignment': Alignment(horizontal='right', vertical='center')
        }
    
    def apply_branding(self, workbook: openpyxl.Workbook) -> openpyxl.Workbook:
        """Apply complete branding to Excel workbook."""
        self.logger.info("Applying branding to Excel workbook")
        
        try:
            # Apply branding to all worksheets
            for worksheet in workbook.worksheets:
                self._apply_worksheet_branding(worksheet)
            
            # Apply workbook-level properties
            self._apply_workbook_properties(workbook)
            
            self.logger.info("Branding applied successfully to Excel workbook")
            return workbook
            
        except Exception as e:
            self.logger.error(f"Failed to apply branding to Excel workbook: {e}")
            raise
    
    def apply_logo(self, workbook: openpyxl.Workbook, position: Any) -> openpyxl.Workbook:
        """Apply logo to workbook (Excel has limited logo support)."""
        # Excel doesn't have the same logo placement options as Word
        # We can add logos as images to specific cells
        if not self.branding.logo.enabled or not self.branding.logo.logo_path:
            return workbook
        
        try:
            # Add logo to the first worksheet
            if workbook.worksheets:
                worksheet = workbook.worksheets[0]
                self._add_logo_to_worksheet(worksheet)
                
        except Exception as e:
            self.logger.error(f"Failed to add logo to Excel workbook: {e}")
        
        return workbook
    
    def apply_color_scheme(self, workbook: openpyxl.Workbook) -> openpyxl.Workbook:
        """Apply color scheme to workbook."""
        try:
            for worksheet in workbook.worksheets:
                self._apply_worksheet_color_scheme(worksheet)
                
        except Exception as e:
            self.logger.error(f"Failed to apply color scheme: {e}")
        
        return workbook
    
    def apply_fonts(self, workbook: openpyxl.Workbook) -> openpyxl.Workbook:
        """Apply font configuration to workbook."""
        try:
            for worksheet in workbook.worksheets:
                self._apply_worksheet_fonts(worksheet)
                
        except Exception as e:
            self.logger.error(f"Failed to apply fonts: {e}")
        
        return workbook
    
    def apply_conditional_formatting(self, workbook: openpyxl.Workbook, theme_name: str) -> openpyxl.Workbook:
        """Apply conditional formatting theme to workbook."""
        if theme_name not in self.branding.formatting_themes:
            self.logger.warning(f"Formatting theme not found: {theme_name}")
            return workbook
        
        theme = self.branding.formatting_themes[theme_name]
        
        try:
            for worksheet in workbook.worksheets:
                self._apply_worksheet_conditional_formatting(worksheet, theme)
                
        except Exception as e:
            self.logger.error(f"Failed to apply conditional formatting: {e}")
        
        return workbook
    
    def _apply_worksheet_branding(self, worksheet):
        """Apply branding to a single worksheet."""
        # Apply fonts
        self._apply_worksheet_fonts(worksheet)
        
        # Apply color scheme
        self._apply_worksheet_color_scheme(worksheet)
        
        # Apply table styling if data exists
        if worksheet.max_row > 1 and worksheet.max_column > 1:
            self._apply_table_styling(worksheet)
    
    def _apply_workbook_properties(self, workbook: openpyxl.Workbook):
        """Apply workbook-level properties."""
        try:
            # Set workbook properties
            workbook.properties.title = f"{self.branding.company_name} - Cloud BOM Report"
            workbook.properties.creator = self.branding.company_name
            workbook.properties.description = "Cloud Bill of Materials Report"
            
        except Exception as e:
            self.logger.error(f"Failed to set workbook properties: {e}")
    
    def _apply_worksheet_fonts(self, worksheet):
        """Apply font styling to worksheet."""
        try:
            # Apply default font to all cells with data
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                         min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    if cell.value is not None:
                        cell.font = self.styles['table_font']
            
            # Apply header font to first row if it exists
            if worksheet.max_row >= 1:
                for cell in worksheet[1]:
                    if cell.value is not None:
                        cell.font = self.styles['header_font']
                        
        except Exception as e:
            self.logger.error(f"Failed to apply worksheet fonts: {e}")
    
    def _apply_worksheet_color_scheme(self, worksheet):
        """Apply color scheme to worksheet."""
        try:
            # Apply header row styling
            if worksheet.max_row >= 1:
                for cell in worksheet[1]:
                    if cell.value is not None:
                        cell.fill = self.styles['header_fill']
                        cell.font = self.styles['header_font']
            
            # Apply alternating row colors
            for row_num in range(2, worksheet.max_row + 1):
                if row_num % 2 == 0:  # Even rows
                    for col_num in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        if cell.value is not None:
                            cell.fill = self.styles['alt_row_fill']
                            
        except Exception as e:
            self.logger.error(f"Failed to apply worksheet color scheme: {e}")
    
    def _apply_table_styling(self, worksheet):
        """Apply professional table styling to worksheet."""
        try:
            # Apply borders to data range
            data_range = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
            
            for row in worksheet[data_range]:
                for cell in row:
                    if cell.value is not None:
                        cell.border = self.styles['thin_border']
                        cell.alignment = self.styles['left_alignment']
            
            # Apply special styling to header row
            if worksheet.max_row >= 1:
                for cell in worksheet[1]:
                    if cell.value is not None:
                        cell.border = self.styles['thick_border']
                        cell.alignment = self.styles['center_alignment']
            
            # Auto-adjust column widths
            self._auto_adjust_column_widths(worksheet)
            
        except Exception as e:
            self.logger.error(f"Failed to apply table styling: {e}")
    
    def _apply_worksheet_conditional_formatting(self, worksheet, theme: ConditionalFormattingTheme):
        """Apply conditional formatting to worksheet."""
        try:
            # Find data range
            if worksheet.max_row <= 1:
                return
            
            data_range = f"A2:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
            
            # Apply compliance status formatting
            compliant_rule = CellIsRule(
                operator='containsText',
                formula=['compliant'],
                fill=self.styles['compliant_fill'],
                font=Font(color=self._hex_to_openpyxl_color(self.branding.colors.compliant))
            )
            worksheet.conditional_formatting.add(data_range, compliant_rule)
            
            non_compliant_rule = CellIsRule(
                operator='containsText',
                formula=['non-compliant'],
                fill=self.styles['non_compliant_fill'],
                font=Font(color=self._hex_to_openpyxl_color(self.branding.colors.non_compliant))
            )
            worksheet.conditional_formatting.add(data_range, non_compliant_rule)
            
            # Apply risk level formatting
            high_risk_rule = CellIsRule(
                operator='containsText',
                formula=['high risk'],
                fill=PatternFill(
                    start_color=self._hex_to_openpyxl_color(ColorUtilities.lighten_color(self.branding.colors.danger, 0.8)),
                    end_color=self._hex_to_openpyxl_color(ColorUtilities.lighten_color(self.branding.colors.danger, 0.8)),
                    fill_type='solid'
                ),
                font=Font(color=self._hex_to_openpyxl_color(self.branding.colors.danger), bold=True)
            )
            worksheet.conditional_formatting.add(data_range, high_risk_rule)
            
        except Exception as e:
            self.logger.error(f"Failed to apply conditional formatting: {e}")
    
    def _add_logo_to_worksheet(self, worksheet):
        """Add logo to worksheet."""
        try:
            from openpyxl.drawing.image import Image
            
            # Create image object
            img = Image(self.branding.logo.logo_path)
            
            # Resize image
            img.width = int(self.branding.logo.size[0] * 72)  # Convert inches to pixels
            img.height = int(self.branding.logo.size[1] * 72)
            
            # Add to worksheet (top-left corner)
            worksheet.add_image(img, 'A1')
            
            # Adjust row height to accommodate logo
            worksheet.row_dimensions[1].height = img.height * 0.75
            
        except Exception as e:
            self.logger.error(f"Failed to add logo to worksheet: {e}")
    
    def _auto_adjust_column_widths(self, worksheet):
        """Auto-adjust column widths based on content."""
        try:
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    if cell.value is not None:
                        try:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass
                
                # Set column width with some padding
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            self.logger.error(f"Failed to auto-adjust column widths: {e}")
    
    def _hex_to_openpyxl_color(self, hex_color: str) -> str:
        """Convert hex color to openpyxl color format."""
        return hex_color.lstrip('#').upper()
    
    def create_branded_chart(self, worksheet, chart_type: str, data_range: str, title: str) -> Any:
        """Create a chart with branding applied."""
        try:
            # Create chart based on type
            if chart_type.lower() == 'bar':
                chart = BarChart()
            elif chart_type.lower() == 'pie':
                chart = PieChart()
            elif chart_type.lower() == 'line':
                chart = LineChart()
            else:
                chart = BarChart()  # Default
            
            # Set chart title
            chart.title = title
            
            # Apply brand colors to chart
            self._apply_chart_branding(chart)
            
            # Add data
            data = openpyxl.chart.Reference(worksheet, range_string=data_range)
            chart.add_data(data, titles_from_data=True)
            
            return chart
            
        except Exception as e:
            self.logger.error(f"Failed to create branded chart: {e}")
            return None
    
    def _apply_chart_branding(self, chart):
        """Apply branding to chart."""
        try:
            # Set chart colors using brand color scheme
            if hasattr(chart, 'series') and chart.series:
                for i, series in enumerate(chart.series):
                    color_index = i % len(self.branding.colors.chart_colors)
                    color = self.branding.colors.chart_colors[color_index]
                    
                    # Apply color to series (this varies by chart type)
                    if hasattr(series, 'graphicalProperties'):
                        series.graphicalProperties.solidFill = color.lstrip('#')
                        
        except Exception as e:
            self.logger.error(f"Failed to apply chart branding: {e}")
    
    def create_branded_table(self, worksheet, data: List[List[str]], headers: List[str], start_cell: str = "A1"):
        """Create a table with branding applied."""
        try:
            # Parse start cell
            from openpyxl.utils import coordinate_from_string, column_index_from_string
            col, row = coordinate_from_string(start_cell)
            start_col = column_index_from_string(col)
            start_row = row
            
            # Add headers
            for i, header in enumerate(headers):
                cell = worksheet.cell(row=start_row, column=start_col + i)
                cell.value = header
                cell.font = self.styles['header_font']
                cell.fill = self.styles['header_fill']
                cell.border = self.styles['thick_border']
                cell.alignment = self.styles['center_alignment']
            
            # Add data
            for row_idx, row_data in enumerate(data):
                for col_idx, cell_data in enumerate(row_data):
                    cell = worksheet.cell(row=start_row + row_idx + 1, column=start_col + col_idx)
                    cell.value = cell_data
                    cell.font = self.styles['table_font']
                    cell.border = self.styles['thin_border']
                    cell.alignment = self.styles['left_alignment']
                    
                    # Apply alternating row colors
                    if row_idx % 2 == 1:  # Odd rows (0-indexed)
                        cell.fill = self.styles['alt_row_fill']
            
            # Auto-adjust column widths
            end_col = start_col + len(headers) - 1
            for col in range(start_col, end_col + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                
                for row in range(start_row, start_row + len(data) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            self.logger.error(f"Failed to create branded table: {e}")
    
    def apply_freeze_panes(self, worksheet, cell: str = "A2"):
        """Apply freeze panes to worksheet."""
        try:
            worksheet.freeze_panes = cell
        except Exception as e:
            self.logger.error(f"Failed to apply freeze panes: {e}")
    
    def apply_auto_filter(self, worksheet, data_range: Optional[str] = None):
        """Apply auto filter to worksheet."""
        try:
            if data_range:
                worksheet.auto_filter.ref = data_range
            else:
                # Auto-detect data range
                if worksheet.max_row > 1 and worksheet.max_column > 1:
                    data_range = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
                    worksheet.auto_filter.ref = data_range
                    
        except Exception as e:
            self.logger.error(f"Failed to apply auto filter: {e}")


# Factory function
def create_excel_branding_applicator(branding_config: AdvancedBrandingConfig) -> ExcelBrandingApplicator:
    """Create an ExcelBrandingApplicator instance."""
    return ExcelBrandingApplicator(branding_config)