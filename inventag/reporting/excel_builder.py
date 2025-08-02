#!/usr/bin/env python3
"""
Excel Workbook Builder

Excel document builder for professional BOM generation.
Enhanced from existing bom_converter.py Excel generation with advanced features.

Features:
- Excel workbook generation with multiple service-specific sheets and formatting
- Executive summary dashboard with compliance metrics and charts
- Network analysis sheet with VPC/subnet utilization charts and capacity planning
- Security analysis sheet with risk assessment tables and recommendations
- Conditional formatting for compliance status highlighting and visual indicators
"""

import logging
import os
from typing import Dict, List, Any, Optional, Set
from datetime import datetime, timezone
from collections import defaultdict

from .document_generator import DocumentBuilder, DocumentGenerationResult, BOMData

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
    from openpyxl.styles.differential import DifferentialStyle
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    # Create placeholders for type hints when openpyxl is not available
    Workbook = Any


class ExcelWorkbookBuilder(DocumentBuilder):
    """
    Excel workbook builder for professional BOM generation.
    
    Enhanced from existing bom_converter.py Excel generation with:
    - Excel workbook generation with multiple service-specific sheets and formatting
    - Executive summary dashboard with compliance metrics and charts
    - Network analysis sheet with VPC/subnet utilization charts and capacity planning
    - Security analysis sheet with risk assessment tables and recommendations
    - Conditional formatting for compliance status highlighting and visual indicators
    """
    
    def __init__(self, config):
        super().__init__(config)
        
        # Excel styling configuration
        self.styles = self._initialize_styles()
        
    def _ensure_argb_color(self, color: str) -> str:
        """Ensure color is in ARGB format for openpyxl."""
        if not color:
            return "FF000000"  # Default black with full alpha
            
        # Remove any # prefix
        color = color.lstrip('#')
        
        # If it's already 8 characters (ARGB), return as is
        if len(color) == 8:
            return color.upper()
            
        # If it's 6 characters (RGB), add full alpha
        if len(color) == 6:
            return f"FF{color.upper()}"
            
        # If it's 3 characters, expand to 6 then add alpha
        if len(color) == 3:
            expanded = ''.join([c*2 for c in color])
            return f"FF{expanded.upper()}"
            
        # Default fallback
        return "FF000000"
        
    def can_handle_format(self, format_type: str) -> bool:
        """Check if this builder can handle Excel format."""
        return format_type.lower() == "excel"
        
    def validate_dependencies(self) -> List[str]:
        """Validate Excel dependencies."""
        if not OPENPYXL_AVAILABLE:
            return ["openpyxl library not available - install with: pip install openpyxl"]
        return []
        
    def _initialize_styles(self) -> Dict[str, Any]:
        """Initialize Excel styling configuration."""
        if not OPENPYXL_AVAILABLE:
            return {}
            
        return {
            "header_font": Font(bold=True, color="FFFFFF"),
            "header_fill": PatternFill(
                start_color=self._ensure_argb_color(self.config.branding.color_scheme.get("primary", "366092")),
                end_color=self._ensure_argb_color(self.config.branding.color_scheme.get("primary", "366092")),
                fill_type="solid"
            ),
            "subheader_font": Font(bold=True, color="000000"),
            "subheader_fill": PatternFill(
                start_color="E7E6E6",
                end_color="E7E6E6", 
                fill_type="solid"
            ),
            "compliant_fill": PatternFill(
                start_color=self._ensure_argb_color(self.config.branding.color_scheme.get("accent", "70AD47")),
                end_color=self._ensure_argb_color(self.config.branding.color_scheme.get("accent", "70AD47")),
                fill_type="solid"
            ),
            "non_compliant_fill": PatternFill(
                start_color=self._ensure_argb_color(self.config.branding.color_scheme.get("danger", "C5504B")),
                end_color=self._ensure_argb_color(self.config.branding.color_scheme.get("danger", "C5504B")),
                fill_type="solid"
            ),
            "warning_fill": PatternFill(
                start_color=self._ensure_argb_color(self.config.branding.color_scheme.get("warning", "FFC000")),
                end_color=self._ensure_argb_color(self.config.branding.color_scheme.get("warning", "FFC000")),
                fill_type="solid"
            ),
            "border": Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            "center_alignment": Alignment(horizontal="center", vertical="center"),
            "left_alignment": Alignment(horizontal="left", vertical="center")
        }
        
    def generate_document(self, bom_data: BOMData, output_path: str) -> DocumentGenerationResult:
        """
        Generate Excel workbook with multiple sheets and advanced formatting.
        
        Enhanced from bom_converter.py Excel generation with:
        - Excel workbook generation with multiple service-specific sheets and formatting
        - Executive summary dashboard with compliance metrics and charts
        - Network analysis sheet with VPC/subnet utilization charts and capacity planning
        - Security analysis sheet with risk assessment tables and recommendations
        - Conditional formatting for compliance status highlighting and visual indicators
        """
        start_time = datetime.now(timezone.utc)
        
        try:
            if not OPENPYXL_AVAILABLE:
                return DocumentGenerationResult(
                    format_type="excel",
                    filename=os.path.basename(output_path) if output_path else "",
                    success=False,
                    error_message="openpyxl library not available"
                )
                
            self.logger.info(f"Generating Excel workbook: {output_path}")
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
                
            # Create sheets in order
            self._create_executive_summary_sheet(wb, bom_data)
            self._create_service_specific_sheets(wb, bom_data)
            self._create_network_analysis_sheet(wb, bom_data)
            self._create_security_analysis_sheet(wb, bom_data)
            self._create_compliance_details_sheet(wb, bom_data)
            
            # Apply branding
            self._apply_workbook_branding(wb)
            
            # Save workbook
            wb.save(output_path)
            
            end_time = datetime.now(timezone.utc)
            
            self.logger.info(f"Excel workbook generated successfully: {output_path}")
            
            return DocumentGenerationResult(
                format_type="excel",
                filename=os.path.basename(output_path),
                success=True,
                generation_time_seconds=(end_time - start_time).total_seconds()
            )
            
        except Exception as e:
            self.logger.error(f"Excel generation failed: {e}")
            return DocumentGenerationResult(
                format_type="excel",
                filename=os.path.basename(output_path) if output_path else "",
                success=False,
                error_message=str(e)
            )
            
    def _create_executive_summary_sheet(self, wb: Workbook, bom_data: BOMData):
        """
        Create executive summary dashboard with compliance metrics and charts.
        
        Enhanced executive summary with:
        - Compliance overview and key metrics
        - Service breakdown with resource counts
        - Compliance status charts
        - Key findings and recommendations
        """
        ws = wb.create_sheet("Executive Summary", 0)
        
        # Title and metadata
        self._add_sheet_title(ws, "Executive Summary Dashboard", 1)
        self._add_metadata_section(ws, bom_data, 3)
        
        # Compliance overview section
        self._add_compliance_overview(ws, bom_data, 8)
        
        # Service breakdown section
        self._add_service_breakdown(ws, bom_data, 15)
        
        # Charts section
        self._add_compliance_charts(ws, bom_data, 25)
        
        # Key findings section
        self._add_key_findings(ws, bom_data, 35)
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
        
    def _create_service_specific_sheets(self, wb: Workbook, bom_data: BOMData):
        """
        Create separate sheets for each AWS service with detailed resource information.
        
        Enhanced service sheets with:
        - Service-specific resource tables with custom descriptions and formatting
        - Conditional formatting for compliance status
        - Service-specific attributes and metadata
        """
        # Group resources by service
        services = defaultdict(list)
        for resource in bom_data.resources:
            service = resource.get("service", "Unknown")
            services[service].append(resource)
            
        # Create sheet for each service
        for service, resources in sorted(services.items()):
            self._create_single_service_sheet(wb, service, resources, bom_data)
            
    def _create_single_service_sheet(self, wb: Workbook, service: str, resources: List[Dict], bom_data: BOMData):
        """Create a sheet for a specific service."""
        # Sanitize sheet name (Excel has restrictions)
        sheet_name = service[:31]  # Excel sheet names are limited to 31 characters
        ws = wb.create_sheet(sheet_name)
        
        # Sheet title
        self._add_sheet_title(ws, f"{service} Resources", 1)
        
        # Service summary
        self._add_service_summary(ws, service, resources, 3)
        
        # Resource table
        self._add_service_resource_table(ws, resources, 8)
        
        # Apply conditional formatting
        self._apply_service_conditional_formatting(ws, len(resources))
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
        
    def _create_network_analysis_sheet(self, wb: Workbook, bom_data: BOMData):
        """
        Create network analysis sheet with VPC/subnet utilization and capacity planning.
        
        Enhanced network analysis with:
        - VPC/subnet utilization charts and capacity planning
        - CIDR utilization details and diagrams
        - Network resource mapping
        """
        ws = wb.create_sheet("Network Analysis")
        
        # Sheet title
        self._add_sheet_title(ws, "Network Analysis & Capacity Planning", 1)
        
        # Network overview
        self._add_network_overview(ws, bom_data, 3)
        
        # VPC utilization table
        self._add_vpc_utilization_table(ws, bom_data, 10)
        
        # Network resource mapping
        self._add_network_resource_mapping(ws, bom_data, 20)
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
        
    def _create_security_analysis_sheet(self, wb: Workbook, bom_data: BOMData):
        """
        Create security analysis sheet with risk assessment and recommendations.
        
        Enhanced security analysis with:
        - Risk assessment tables and recommendations
        - Security group analysis with overly permissive rules
        - Security findings and remediation guidance
        """
        ws = wb.create_sheet("Security Analysis")
        
        # Sheet title
        self._add_sheet_title(ws, "Security Analysis & Risk Assessment", 1)
        
        # Security overview
        self._add_security_overview(ws, bom_data, 3)
        
        # Risk assessment table
        self._add_risk_assessment_table(ws, bom_data, 10)
        
        # Security recommendations
        self._add_security_recommendations(ws, bom_data, 20)
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
        
    def _create_compliance_details_sheet(self, wb: Workbook, bom_data: BOMData):
        """Create detailed compliance information sheet."""
        ws = wb.create_sheet("Compliance Details")
        
        # Sheet title
        self._add_sheet_title(ws, "Detailed Compliance Information", 1)
        
        # Compliance breakdown by service
        self._add_compliance_breakdown(ws, bom_data, 3)
        
        # Non-compliant resources details
        self._add_non_compliant_details(ws, bom_data, 15)
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
        
    def _add_sheet_title(self, ws, title: str, row: int):
        """Add formatted title to sheet."""
        ws.merge_cells(f"A{row}:H{row}")
        cell = ws[f"A{row}"]
        cell.value = title
        cell.font = Font(size=16, bold=True, color="000000")
        cell.alignment = self.styles["center_alignment"]
        cell.fill = self.styles["subheader_fill"]
        
    def _add_metadata_section(self, ws, bom_data: BOMData, start_row: int):
        """Add metadata section to sheet."""
        metadata = bom_data.generation_metadata
        
        ws[f"A{start_row}"] = "Report Generated:"
        ws[f"B{start_row}"] = metadata.get("generated_at", "Unknown")
        
        ws[f"A{start_row + 1}"] = "Total Resources:"
        ws[f"B{start_row + 1}"] = metadata.get("total_resources", 0)
        
        ws[f"A{start_row + 2}"] = "Processing Time:"
        ws[f"B{start_row + 2}"] = f"{metadata.get('processing_time_seconds', 0):.2f} seconds"
        
        # Apply formatting
        for row in range(start_row, start_row + 3):
            ws[f"A{row}"].font = Font(bold=True)
            
    def _add_compliance_overview(self, ws, bom_data: BOMData, start_row: int):
        """Add compliance overview section."""
        compliance = bom_data.compliance_summary
        
        # Section header
        ws.merge_cells(f"A{start_row}:D{start_row}")
        ws[f"A{start_row}"] = "Compliance Overview"
        ws[f"A{start_row}"].font = self.styles["subheader_font"]
        ws[f"A{start_row}"].fill = self.styles["subheader_fill"]
        
        # Compliance metrics
        row = start_row + 2
        ws[f"A{row}"] = "Total Resources"
        ws[f"B{row}"] = compliance.get("total_resources", 0)
        
        ws[f"A{row + 1}"] = "Compliant Resources"
        ws[f"B{row + 1}"] = compliance.get("compliant_resources", 0)
        ws[f"B{row + 1}"].fill = self.styles["compliant_fill"]
        
        ws[f"A{row + 2}"] = "Non-Compliant Resources"
        ws[f"B{row + 2}"] = compliance.get("non_compliant_resources", 0)
        ws[f"B{row + 2}"].fill = self.styles["non_compliant_fill"]
        
        ws[f"A{row + 3}"] = "Compliance Percentage"
        compliance_pct = compliance.get("compliance_percentage", 0)
        ws[f"B{row + 3}"] = f"{compliance_pct:.1f}%"
        
        # Apply formatting
        for r in range(row, row + 4):
            ws[f"A{r}"].font = Font(bold=True)
            
    def _add_service_breakdown(self, ws, bom_data: BOMData, start_row: int):
        """Add service breakdown section."""
        # Count resources by service
        services = defaultdict(int)
        for resource in bom_data.resources:
            service = resource.get("service", "Unknown")
            services[service] += 1
            
        # Section header
        ws.merge_cells(f"A{start_row}:C{start_row}")
        ws[f"A{start_row}"] = "Service Breakdown"
        ws[f"A{start_row}"].font = self.styles["subheader_font"]
        ws[f"A{start_row}"].fill = self.styles["subheader_fill"]
        
        # Table headers
        row = start_row + 2
        ws[f"A{row}"] = "Service"
        ws[f"B{row}"] = "Resource Count"
        ws[f"C{row}"] = "Percentage"
        
        # Apply header formatting
        for col in ["A", "B", "C"]:
            cell = ws[f"{col}{row}"]
            cell.font = self.styles["header_font"]
            cell.fill = self.styles["header_fill"]
            cell.alignment = self.styles["center_alignment"]
            
        # Service data
        total_resources = sum(services.values())
        row += 1
        
        for service, count in sorted(services.items()):
            percentage = (count / total_resources * 100) if total_resources > 0 else 0
            
            ws[f"A{row}"] = service
            ws[f"B{row}"] = count
            ws[f"C{row}"] = f"{percentage:.1f}%"
            
            # Apply borders
            for col in ["A", "B", "C"]:
                ws[f"{col}{row}"].border = self.styles["border"]
                
            row += 1
            
    def _add_compliance_charts(self, ws, bom_data: BOMData, start_row: int):
        """Add compliance charts to the sheet."""
        if not OPENPYXL_AVAILABLE:
            return
            
        compliance = bom_data.compliance_summary
        
        # Create pie chart for compliance status
        chart = PieChart()
        chart.title = "Compliance Status Distribution"
        
        # Add data for chart (simplified for now)
        ws[f"F{start_row}"] = "Compliant"
        ws[f"G{start_row}"] = compliance.get("compliant_resources", 0)
        
        ws[f"F{start_row + 1}"] = "Non-Compliant"
        ws[f"G{start_row + 1}"] = compliance.get("non_compliant_resources", 0)
        
        # Configure chart
        data = Reference(ws, min_col=7, min_row=start_row, max_row=start_row + 1)
        labels = Reference(ws, min_col=6, min_row=start_row, max_row=start_row + 1)
        
        chart.add_data(data)
        chart.set_categories(labels)
        chart.height = 10
        chart.width = 15
        
        # Add chart to sheet
        ws.add_chart(chart, f"F{start_row + 3}")
        
    def _add_key_findings(self, ws, bom_data: BOMData, start_row: int):
        """Add key findings and recommendations."""
        # Section header
        ws.merge_cells(f"A{start_row}:H{start_row}")
        ws[f"A{start_row}"] = "Key Findings & Recommendations"
        ws[f"A{start_row}"].font = self.styles["subheader_font"]
        ws[f"A{start_row}"].fill = self.styles["subheader_fill"]
        
        # Generate findings based on data
        findings = self._generate_key_findings(bom_data)
        
        row = start_row + 2
        for finding in findings:
            ws[f"A{row}"] = f"• {finding}"
            ws[f"A{row}"].alignment = self.styles["left_alignment"]
            row += 1
            
    def _generate_key_findings(self, bom_data: BOMData) -> List[str]:
        """Generate key findings based on BOM data."""
        findings = []
        
        compliance = bom_data.compliance_summary
        total_resources = compliance.get("total_resources", 0)
        compliant_resources = compliance.get("compliant_resources", 0)
        
        if total_resources > 0:
            compliance_pct = (compliant_resources / total_resources) * 100
            
            if compliance_pct >= 90:
                findings.append(f"Excellent compliance rate of {compliance_pct:.1f}% - organization follows best practices")
            elif compliance_pct >= 70:
                findings.append(f"Good compliance rate of {compliance_pct:.1f}% - minor improvements needed")
            else:
                findings.append(f"Compliance rate of {compliance_pct:.1f}% requires immediate attention")
                
        # Network findings
        network_analysis = bom_data.network_analysis
        if network_analysis:
            total_vpcs = network_analysis.get("total_vpcs", 0)
            if total_vpcs > 0:
                findings.append(f"Infrastructure spans {total_vpcs} VPCs - review for optimization opportunities")
                
        # Security findings
        security_analysis = bom_data.security_analysis
        if security_analysis:
            high_risk_rules = security_analysis.get("high_risk_rules", 0)
            if high_risk_rules > 0:
                findings.append(f"Found {high_risk_rules} high-risk security rules - immediate review recommended")
                
        return findings
        
    def _add_service_summary(self, ws, service: str, resources: List[Dict], start_row: int):
        """Add service summary information."""
        ws[f"A{start_row}"] = f"Service: {service}"
        ws[f"A{start_row}"].font = Font(bold=True, size=12)
        
        ws[f"A{start_row + 1}"] = f"Total Resources: {len(resources)}"
        
        # Count compliance status
        compliant = sum(1 for r in resources if r.get("compliance_status") == "compliant")
        non_compliant = len(resources) - compliant
        
        ws[f"A{start_row + 2}"] = f"Compliant: {compliant}"
        ws[f"A{start_row + 2}"].fill = self.styles["compliant_fill"]
        
        ws[f"A{start_row + 3}"] = f"Non-Compliant: {non_compliant}"
        ws[f"A{start_row + 3}"].fill = self.styles["non_compliant_fill"]
        
    def _add_service_resource_table(self, ws, resources: List[Dict], start_row: int):
        """Add resource table for a service."""
        if not resources:
            return
            
        # Get all unique headers for this service
        headers = set()
        for resource in resources:
            flattened = self._flatten_dict(resource)
            headers.update(flattened.keys())
            
        sorted_headers = sorted(headers)
        
        # Write headers
        for col, header in enumerate(sorted_headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = self.styles["header_font"]
            cell.fill = self.styles["header_fill"]
            cell.alignment = self.styles["center_alignment"]
            cell.border = self.styles["border"]
            
        # Write data
        for row, resource in enumerate(resources, start_row + 1):
            flattened = self._flatten_dict(resource)
            for col, header in enumerate(sorted_headers, 1):
                value = flattened.get(header, "")
                cell = ws.cell(row=row, column=col, value=str(value) if value is not None else "")
                cell.border = self.styles["border"]
                cell.alignment = self.styles["left_alignment"]
                
    def _apply_service_conditional_formatting(self, ws, num_resources: int):
        """Apply conditional formatting to service sheet."""
        if not OPENPYXL_AVAILABLE or num_resources == 0:
            return
            
        # Find compliance_status column
        compliance_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=8, column=col).value == "compliance_status":  # Assuming headers start at row 8
                compliance_col = col
                break
                
        if compliance_col:
            # Apply conditional formatting for compliance status
            compliant_rule = CellIsRule(
                operator='equal',
                formula=['"compliant"'],
                fill=self.styles["compliant_fill"]
            )
            
            non_compliant_rule = CellIsRule(
                operator='equal', 
                formula=['"non_compliant"'],
                fill=self.styles["non_compliant_fill"]
            )
            
            # Apply to compliance column
            col_letter = get_column_letter(compliance_col)
            range_str = f"{col_letter}9:{col_letter}{8 + num_resources}"
            
            ws.conditional_formatting.add(range_str, compliant_rule)
            ws.conditional_formatting.add(range_str, non_compliant_rule)
            
    def _add_network_overview(self, ws, bom_data: BOMData, start_row: int):
        """Add network overview section."""
        network_analysis = bom_data.network_analysis
        
        ws[f"A{start_row}"] = "Network Overview"
        ws[f"A{start_row}"].font = self.styles["subheader_font"]
        ws[f"A{start_row}"].fill = self.styles["subheader_fill"]
        
        row = start_row + 2
        ws[f"A{row}"] = "Total VPCs:"
        ws[f"B{row}"] = network_analysis.get("total_vpcs", 0)
        
        ws[f"A{row + 1}"] = "Total Subnets:"
        ws[f"B{row + 1}"] = network_analysis.get("total_subnets", 0)
        
        # Apply formatting
        for r in range(row, row + 2):
            ws[f"A{r}"].font = Font(bold=True)
            
    def _add_vpc_utilization_table(self, ws, bom_data: BOMData, start_row: int):
        """Add VPC utilization table."""
        network_analysis = bom_data.network_analysis
        vpc_utilization = network_analysis.get("vpc_utilization", {})
        
        if not vpc_utilization:
            return
            
        # Table header
        ws[f"A{start_row}"] = "VPC Utilization Analysis"
        ws[f"A{start_row}"].font = self.styles["subheader_font"]
        ws[f"A{start_row}"].fill = self.styles["subheader_fill"]
        
        # Column headers
        headers = ["VPC ID", "VPC Name", "CIDR Block", "Utilization %", "Available IPs"]
        row = start_row + 2
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles["header_font"]
            cell.fill = self.styles["header_fill"]
            cell.alignment = self.styles["center_alignment"]
            cell.border = self.styles["border"]
            
        # VPC data
        row += 1
        for vpc_id, vpc_data in vpc_utilization.items():
            ws.cell(row=row, column=1, value=vpc_id).border = self.styles["border"]
            ws.cell(row=row, column=2, value=vpc_data.get("name", "")).border = self.styles["border"]
            ws.cell(row=row, column=3, value=vpc_data.get("cidr_block", "")).border = self.styles["border"]
            ws.cell(row=row, column=4, value=f"{vpc_data.get('utilization_percentage', 0):.1f}%").border = self.styles["border"]
            ws.cell(row=row, column=5, value=vpc_data.get("available_ips", 0)).border = self.styles["border"]
            row += 1
            
    def _add_network_resource_mapping(self, ws, bom_data: BOMData, start_row: int):
        """Add network resource mapping section."""
        # Count resources by VPC
        vpc_resources = defaultdict(int)
        for resource in bom_data.resources:
            vpc_id = resource.get("vpc_id")
            if vpc_id:
                vpc_resources[vpc_id] += 1
                
        if not vpc_resources:
            return
            
        ws[f"A{start_row}"] = "Resources by VPC"
        ws[f"A{start_row}"].font = self.styles["subheader_font"]
        ws[f"A{start_row}"].fill = self.styles["subheader_fill"]
        
        row = start_row + 2
        ws[f"A{row}"] = "VPC ID"
        ws[f"B{row}"] = "Resource Count"
        
        # Apply header formatting
        for col in ["A", "B"]:
            cell = ws[f"{col}{row}"]
            cell.font = self.styles["header_font"]
            cell.fill = self.styles["header_fill"]
            cell.border = self.styles["border"]
            
        row += 1
        for vpc_id, count in sorted(vpc_resources.items()):
            ws.cell(row=row, column=1, value=vpc_id).border = self.styles["border"]
            ws.cell(row=row, column=2, value=count).border = self.styles["border"]
            row += 1
            
    def _add_security_overview(self, ws, bom_data: BOMData, start_row: int):
        """Add security overview section."""
        security_analysis = bom_data.security_analysis
        
        ws[f"A{start_row}"] = "Security Overview"
        ws[f"A{start_row}"].font = self.styles["subheader_font"]
        ws[f"A{start_row}"].fill = self.styles["subheader_fill"]
        
        row = start_row + 2
        ws[f"A{row}"] = "Total Security Groups:"
        ws[f"B{row}"] = security_analysis.get("total_security_groups", 0)
        
        ws[f"A{row + 1}"] = "High Risk Rules:"
        ws[f"B{row + 1}"] = security_analysis.get("high_risk_rules", 0)
        ws[f"B{row + 1}"].fill = self.styles["warning_fill"]
        
        # Apply formatting
        for r in range(row, row + 2):
            ws[f"A{r}"].font = Font(bold=True)
            
    def _add_risk_assessment_table(self, ws, bom_data: BOMData, start_row: int):
        """Add risk assessment table."""
        security_analysis = bom_data.security_analysis
        overly_permissive_rules = security_analysis.get("overly_permissive_rules", [])
        
        if not overly_permissive_rules:
            return
            
        ws[f"A{start_row}"] = "High Risk Security Rules"
        ws[f"A{start_row}"].font = self.styles["subheader_font"]
        ws[f"A{start_row}"].fill = self.styles["subheader_fill"]
        
        # Column headers
        headers = ["Security Group", "Rule", "Risk Level", "Recommendation"]
        row = start_row + 2
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles["header_font"]
            cell.fill = self.styles["header_fill"]
            cell.alignment = self.styles["center_alignment"]
            cell.border = self.styles["border"]
            
        # Risk data
        row += 1
        for rule in overly_permissive_rules:
            ws.cell(row=row, column=1, value=rule.get("group_id", "")).border = self.styles["border"]
            ws.cell(row=row, column=2, value=rule.get("rule", "")).border = self.styles["border"]
            
            risk_cell = ws.cell(row=row, column=3, value=rule.get("risk_level", ""))
            risk_cell.border = self.styles["border"]
            if rule.get("risk_level") == "high":
                risk_cell.fill = self.styles["non_compliant_fill"]
                
            ws.cell(row=row, column=4, value="Restrict source to specific CIDR blocks").border = self.styles["border"]
            row += 1
            
    def _add_security_recommendations(self, ws, bom_data: BOMData, start_row: int):
        """Add security recommendations."""
        recommendations = [
            "Review and restrict overly permissive security group rules",
            "Implement least privilege access principles",
            "Regular security group audits and cleanup",
            "Use AWS Config rules for continuous compliance monitoring",
            "Enable VPC Flow Logs for network traffic analysis"
        ]
        
        ws[f"A{start_row}"] = "Security Recommendations"
        ws[f"A{start_row}"].font = self.styles["subheader_font"]
        ws[f"A{start_row}"].fill = self.styles["subheader_fill"]
        
        row = start_row + 2
        for recommendation in recommendations:
            ws[f"A{row}"] = f"• {recommendation}"
            ws[f"A{row}"].alignment = self.styles["left_alignment"]
            row += 1
            
    def _add_compliance_breakdown(self, ws, bom_data: BOMData, start_row: int):
        """Add compliance breakdown by service."""
        # Count compliance by service
        service_compliance = defaultdict(lambda: {"compliant": 0, "non_compliant": 0})
        
        for resource in bom_data.resources:
            service = resource.get("service", "Unknown")
            status = resource.get("compliance_status", "unknown")
            
            if status == "compliant":
                service_compliance[service]["compliant"] += 1
            elif status == "non_compliant":
                service_compliance[service]["non_compliant"] += 1
                
        ws[f"A{start_row}"] = "Compliance Breakdown by Service"
        ws[f"A{start_row}"].font = self.styles["subheader_font"]
        ws[f"A{start_row}"].fill = self.styles["subheader_fill"]
        
        # Table headers
        headers = ["Service", "Compliant", "Non-Compliant", "Total", "Compliance %"]
        row = start_row + 2
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles["header_font"]
            cell.fill = self.styles["header_fill"]
            cell.alignment = self.styles["center_alignment"]
            cell.border = self.styles["border"]
            
        # Service data
        row += 1
        for service, counts in sorted(service_compliance.items()):
            compliant = counts["compliant"]
            non_compliant = counts["non_compliant"]
            total = compliant + non_compliant
            compliance_pct = (compliant / total * 100) if total > 0 else 0
            
            ws.cell(row=row, column=1, value=service).border = self.styles["border"]
            
            compliant_cell = ws.cell(row=row, column=2, value=compliant)
            compliant_cell.border = self.styles["border"]
            compliant_cell.fill = self.styles["compliant_fill"]
            
            non_compliant_cell = ws.cell(row=row, column=3, value=non_compliant)
            non_compliant_cell.border = self.styles["border"]
            if non_compliant > 0:
                non_compliant_cell.fill = self.styles["non_compliant_fill"]
                
            ws.cell(row=row, column=4, value=total).border = self.styles["border"]
            ws.cell(row=row, column=5, value=f"{compliance_pct:.1f}%").border = self.styles["border"]
            
            row += 1
            
    def _add_non_compliant_details(self, ws, bom_data: BOMData, start_row: int):
        """Add details of non-compliant resources."""
        non_compliant_resources = [
            r for r in bom_data.resources 
            if r.get("compliance_status") == "non_compliant"
        ]
        
        if not non_compliant_resources:
            ws[f"A{start_row}"] = "No non-compliant resources found"
            return
            
        ws[f"A{start_row}"] = "Non-Compliant Resources Details"
        ws[f"A{start_row}"].font = self.styles["subheader_font"]
        ws[f"A{start_row}"].fill = self.styles["subheader_fill"]
        
        # Table headers
        headers = ["Service", "Type", "ID", "Name", "Region", "Issues"]
        row = start_row + 2
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles["header_font"]
            cell.fill = self.styles["header_fill"]
            cell.alignment = self.styles["center_alignment"]
            cell.border = self.styles["border"]
            
        # Resource data
        row += 1
        for resource in non_compliant_resources:
            ws.cell(row=row, column=1, value=resource.get("service", "")).border = self.styles["border"]
            ws.cell(row=row, column=2, value=resource.get("type", "")).border = self.styles["border"]
            ws.cell(row=row, column=3, value=resource.get("id", "")).border = self.styles["border"]
            ws.cell(row=row, column=4, value=resource.get("name", "")).border = self.styles["border"]
            ws.cell(row=row, column=5, value=resource.get("region", "")).border = self.styles["border"]
            ws.cell(row=row, column=6, value="Missing required tags").border = self.styles["border"]
            
            # Highlight row
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
                )
                
            row += 1
            
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
                    
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
            
    def _apply_workbook_branding(self, wb: Workbook):
        """Apply branding to the entire workbook."""
        # Set workbook properties
        wb.properties.title = f"{self.config.branding.company_name} - Cloud BOM Report"
        wb.properties.creator = "InvenTag Cloud BOM Generator"
        wb.properties.description = "Professional AWS resource inventory and compliance report"
        
    def _flatten_dict(self, d: Dict[str, Any], parent_key: str = "", sep: str = ".") -> Dict[str, Any]:
        """Flatten a nested dictionary."""
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(self._flatten_dict(v, new_key, sep=sep).items())
            elif isinstance(v, list):
                # Convert lists to comma-separated strings
                items.append((new_key, ", ".join(str(item) for item in v)))
            else:
                items.append((new_key, v))
        return dict(items)
        
    def apply_branding(self, workbook: Workbook) -> Workbook:
        """Apply branding configuration to Excel workbook."""
        self._apply_workbook_branding(workbook)
        return workbook