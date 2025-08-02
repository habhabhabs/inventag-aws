#!/usr/bin/env python3
"""
Unit tests for ExcelWorkbookBuilder.

Tests the Excel workbook generation functionality with advanced features.
"""

import unittest
import tempfile
import os
import shutil
from unittest.mock import Mock, patch, MagicMock
from datetime import datetime, timezone

# Import the modules to test
from inventag.reporting.excel_builder import ExcelWorkbookBuilder
from inventag.reporting.document_generator import DocumentConfig, BrandingConfig, DocumentGenerationResult
from inventag.reporting.bom_processor import BOMData

# Check if openpyxl is available
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class TestExcelWorkbookBuilder(unittest.TestCase):
    """Test cases for ExcelWorkbookBuilder."""
    
    def setUp(self):
        """Set up test fixtures."""
        self.temp_dir = tempfile.mkdtemp()
        
        # Create test BOM data
        self.test_bom_data = BOMData(
            resources=[
                {
                    "service": "EC2",
                    "type": "Instance",
                    "id": "i-1234567890abcdef0",
                    "name": "web-server-1",
                    "region": "us-east-1",
                    "compliance_status": "compliant",
                    "tags": {
                        "Name": "web-server-1",
                        "Environment": "production"
                    },
                    "vpc_id": "vpc-12345678"
                },
                {
                    "service": "S3",
                    "type": "Bucket",
                    "id": "test-bucket",
                    "name": "test-bucket",
                    "region": "us-east-1",
                    "compliance_status": "non_compliant",
                    "tags": {
                        "Name": "test-bucket"
                    }
                },
                {
                    "service": "EC2",
                    "type": "Volume",
                    "id": "vol-1234567890abcdef0",
                    "name": "web-server-volume",
                    "region": "us-east-1",
                    "compliance_status": "compliant",
                    "vpc_id": "vpc-12345678"
                }
            ],
            network_analysis={
                "total_vpcs": 1,
                "total_subnets": 2,
                "vpc_utilization": {
                    "vpc-12345678": {
                        "name": "main-vpc",
                        "cidr_block": "10.0.0.0/16",
                        "utilization_percentage": 25.5,
                        "available_ips": 65280
                    }
                }
            },
            security_analysis={
                "total_security_groups": 3,
                "high_risk_rules": 1,
                "overly_permissive_rules": [
                    {
                        "group_id": "sg-12345678",
                        "rule": "0.0.0.0/0:22",
                        "risk_level": "high"
                    }
                ]
            },
            compliance_summary={
                "total_resources": 3,
                "compliant_resources": 2,
                "non_compliant_resources": 1,
                "compliance_percentage": 66.7
            },
            generation_metadata={
                "generated_at": datetime.now(timezone.utc).isoformat(),
                "total_resources": 3,
                "processing_time_seconds": 1.5
            },
            custom_attributes=["inventag:remarks", "inventag:costcenter"]
        )
        
        # Create test configuration
        self.test_config = DocumentConfig(
            output_directory=self.temp_dir,
            branding=BrandingConfig(
                company_name="Test Corporation",
                color_scheme={
                    "primary": "366092",
                    "accent": "70AD47",
                    "danger": "C5504B",
                    "warning": "FFC000"
                }
            )
        )
        
    def tearDown(self):
        """Clean up test fixtures."""
        shutil.rmtree(self.temp_dir, ignore_errors=True)
        
    def test_excel_builder_initialization(self):
        """Test ExcelWorkbookBuilder initialization."""
        builder = ExcelWorkbookBuilder(self.test_config)
        
        self.assertIsInstance(builder, ExcelWorkbookBuilder)
        self.assertEqual(builder.config, self.test_config)
        self.assertIsInstance(builder.styles, dict)
        
    def test_can_handle_format(self):
        """Test format handling capability."""
        builder = ExcelWorkbookBuilder(self.test_config)
        
        self.assertTrue(builder.can_handle_format("excel"))
        self.assertTrue(builder.can_handle_format("EXCEL"))
        self.assertFalse(builder.can_handle_format("word"))
        self.assertFalse(builder.can_handle_format("csv"))
        
    def test_validate_dependencies(self):
        """Test dependency validation."""
        builder = ExcelWorkbookBuilder(self.test_config)
        dependencies = builder.validate_dependencies()
        
        if OPENPYXL_AVAILABLE:
            self.assertEqual(dependencies, [])
        else:
            self.assertIn("openpyxl library not available", dependencies[0])
            
    def test_styles_initialization(self):
        """Test Excel styles initialization."""
        builder = ExcelWorkbookBuilder(self.test_config)
        
        if OPENPYXL_AVAILABLE:
            self.assertIn("header_font", builder.styles)
            self.assertIn("header_fill", builder.styles)
            self.assertIn("compliant_fill", builder.styles)
            self.assertIn("non_compliant_fill", builder.styles)
            self.assertIn("border", builder.styles)
            
    @unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl not available")
    def test_generate_document_success(self):
        """Test successful Excel document generation."""
        builder = ExcelWorkbookBuilder(self.test_config)
        output_path = os.path.join(self.temp_dir, "test_report.xlsx")
        
        result = builder.generate_document(self.test_bom_data, output_path)
        
        self.assertIsInstance(result, DocumentGenerationResult)
        self.assertTrue(result.success)
        self.assertEqual(result.format_type, "excel")
        self.assertEqual(result.filename, "test_report.xlsx")
        self.assertIsNone(result.error_message)
        self.assertGreater(result.generation_time_seconds, 0)
        
        # Verify file was created
        self.assertTrue(os.path.exists(output_path))
        self.assertGreater(os.path.getsize(output_path), 0)
        
    @unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl not available")
    def test_workbook_structure(self):
        """Test Excel workbook structure and sheets."""
        builder = ExcelWorkbookBuilder(self.test_config)
        output_path = os.path.join(self.temp_dir, "structure_test.xlsx")
        
        result = builder.generate_document(self.test_bom_data, output_path)
        self.assertTrue(result.success)
        
        # Load and verify workbook structure
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        
        # Check expected sheets exist
        expected_sheets = [
            "Executive Summary",
            "EC2",  # Service with resources
            "S3",   # Service with resources
            "Network Analysis",
            "Security Analysis",
            "Compliance Details"
        ]
        
        for sheet_name in expected_sheets:
            self.assertIn(sheet_name, wb.sheetnames, f"Sheet '{sheet_name}' not found")
            
        # Verify Executive Summary is first sheet
        self.assertEqual(wb.sheetnames[0], "Executive Summary")
        
    @unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl not available")
    def test_executive_summary_content(self):
        """Test executive summary sheet content."""
        builder = ExcelWorkbookBuilder(self.test_config)
        output_path = os.path.join(self.temp_dir, "summary_test.xlsx")
        
        result = builder.generate_document(self.test_bom_data, output_path)
        self.assertTrue(result.success)
        
        # Load and verify executive summary content
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb["Executive Summary"]
        
        # Check title
        self.assertEqual(ws["A1"].value, "Executive Summary Dashboard")
        
        # Check for compliance overview section
        found_compliance_overview = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "Compliance Overview":
                    found_compliance_overview = True
                    break
            if found_compliance_overview:
                break
                
        self.assertTrue(found_compliance_overview, "Compliance Overview section not found")
        
    @unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl not available")
    def test_service_sheets_content(self):
        """Test service-specific sheets content."""
        builder = ExcelWorkbookBuilder(self.test_config)
        output_path = os.path.join(self.temp_dir, "service_test.xlsx")
        
        result = builder.generate_document(self.test_bom_data, output_path)
        self.assertTrue(result.success)
        
        # Load and verify service sheets
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        
        # Check EC2 sheet (has 2 resources)
        ec2_ws = wb["EC2"]
        self.assertEqual(ec2_ws["A1"].value, "EC2 Resources")
        
        # Check S3 sheet (has 1 resource)
        s3_ws = wb["S3"]
        self.assertEqual(s3_ws["A1"].value, "S3 Resources")
        
    @unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl not available")
    def test_network_analysis_sheet(self):
        """Test network analysis sheet content."""
        builder = ExcelWorkbookBuilder(self.test_config)
        output_path = os.path.join(self.temp_dir, "network_test.xlsx")
        
        result = builder.generate_document(self.test_bom_data, output_path)
        self.assertTrue(result.success)
        
        # Load and verify network analysis sheet
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb["Network Analysis"]
        
        self.assertEqual(ws["A1"].value, "Network Analysis & Capacity Planning")
        
        # Check for network overview
        found_network_overview = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "Network Overview":
                    found_network_overview = True
                    break
            if found_network_overview:
                break
                
        self.assertTrue(found_network_overview, "Network Overview section not found")
        
    @unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl not available")
    def test_security_analysis_sheet(self):
        """Test security analysis sheet content."""
        builder = ExcelWorkbookBuilder(self.test_config)
        output_path = os.path.join(self.temp_dir, "security_test.xlsx")
        
        result = builder.generate_document(self.test_bom_data, output_path)
        self.assertTrue(result.success)
        
        # Load and verify security analysis sheet
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb["Security Analysis"]
        
        self.assertEqual(ws["A1"].value, "Security Analysis & Risk Assessment")
        
        # Check for security overview
        found_security_overview = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "Security Overview":
                    found_security_overview = True
                    break
            if found_security_overview:
                break
                
        self.assertTrue(found_security_overview, "Security Overview section not found")
        
    @unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl not available")
    def test_compliance_details_sheet(self):
        """Test compliance details sheet content."""
        builder = ExcelWorkbookBuilder(self.test_config)
        output_path = os.path.join(self.temp_dir, "compliance_test.xlsx")
        
        result = builder.generate_document(self.test_bom_data, output_path)
        self.assertTrue(result.success)
        
        # Load and verify compliance details sheet
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb["Compliance Details"]
        
        self.assertEqual(ws["A1"].value, "Detailed Compliance Information")
        
    def test_generate_document_without_openpyxl(self):
        """Test document generation when openpyxl is not available."""
        builder = ExcelWorkbookBuilder(self.test_config)
        output_path = os.path.join(self.temp_dir, "no_openpyxl_test.xlsx")
        
        # Mock openpyxl as not available
        with patch('inventag.reporting.excel_builder.OPENPYXL_AVAILABLE', False):
            result = builder.generate_document(self.test_bom_data, output_path)
            
        self.assertFalse(result.success)
        self.assertEqual(result.format_type, "excel")
        self.assertIn("openpyxl library not available", result.error_message)
        
    def test_generate_document_with_exception(self):
        """Test document generation with exception handling."""
        builder = ExcelWorkbookBuilder(self.test_config)
        
        # Create invalid BOM data to trigger an exception during processing
        invalid_bom_data = BOMData(
            resources=[{"invalid": "data"}],  # Missing required fields
            generation_metadata=None  # This should cause an error
        )
        
        output_path = os.path.join(self.temp_dir, "exception_test.xlsx")
        result = builder.generate_document(invalid_bom_data, output_path)
        
        # Should handle the exception gracefully
        self.assertFalse(result.success)
        self.assertEqual(result.format_type, "excel")
        self.assertIsNotNone(result.error_message)
        
    def test_flatten_dict_functionality(self):
        """Test dictionary flattening functionality."""
        builder = ExcelWorkbookBuilder(self.test_config)
        
        test_dict = {
            "simple": "value",
            "nested": {
                "key1": "value1",
                "key2": "value2"
            },
            "list_field": ["item1", "item2", "item3"]
        }
        
        flattened = builder._flatten_dict(test_dict)
        
        self.assertEqual(flattened["simple"], "value")
        self.assertEqual(flattened["nested.key1"], "value1")
        self.assertEqual(flattened["nested.key2"], "value2")
        self.assertEqual(flattened["list_field"], "item1, item2, item3")
        
    def test_generate_key_findings(self):
        """Test key findings generation."""
        builder = ExcelWorkbookBuilder(self.test_config)
        
        findings = builder._generate_key_findings(self.test_bom_data)
        
        self.assertIsInstance(findings, list)
        self.assertGreater(len(findings), 0)
        
        # Should include compliance finding
        compliance_finding = any("compliance rate" in finding.lower() for finding in findings)
        self.assertTrue(compliance_finding, "Compliance finding not generated")
        
        # Should include security finding (we have high risk rules)
        security_finding = any("high-risk security rules" in finding.lower() for finding in findings)
        self.assertTrue(security_finding, "Security finding not generated")
        
    @unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl not available")
    def test_branding_application(self):
        """Test branding application to workbook."""
        builder = ExcelWorkbookBuilder(self.test_config)
        
        from openpyxl import Workbook
        wb = Workbook()
        
        branded_wb = builder.apply_branding(wb)
        
        self.assertEqual(branded_wb.properties.title, "Test Corporation - Cloud BOM Report")
        self.assertEqual(branded_wb.properties.creator, "InvenTag Cloud BOM Generator")
        
    def test_custom_branding_colors(self):
        """Test custom branding colors in styles."""
        custom_config = DocumentConfig(
            branding=BrandingConfig(
                color_scheme={
                    "primary": "FF0000",
                    "accent": "00FF00",
                    "danger": "0000FF",
                    "warning": "FFFF00"
                }
            )
        )
        
        builder = ExcelWorkbookBuilder(custom_config)
        
        if OPENPYXL_AVAILABLE:
            # Check that custom colors are used in styles (openpyxl uses ARGB format)
            self.assertIn("FF0000", builder.styles["header_fill"].start_color.rgb)
            self.assertIn("00FF00", builder.styles["compliant_fill"].start_color.rgb)
            self.assertIn("0000FF", builder.styles["non_compliant_fill"].start_color.rgb)
            self.assertIn("FFFF00", builder.styles["warning_fill"].start_color.rgb)
            
    @unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl not available")
    def test_empty_data_handling(self):
        """Test handling of empty or minimal data."""
        empty_bom_data = BOMData(
            resources=[],
            network_analysis={},
            security_analysis={},
            compliance_summary={},
            generation_metadata={}
        )
        
        builder = ExcelWorkbookBuilder(self.test_config)
        output_path = os.path.join(self.temp_dir, "empty_test.xlsx")
        
        result = builder.generate_document(empty_bom_data, output_path)
        
        # Should still succeed but with empty sheets
        self.assertTrue(result.success)
        self.assertTrue(os.path.exists(output_path))
        
    @unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl not available")
    def test_large_dataset_handling(self):
        """Test handling of larger datasets."""
        # Create larger test dataset
        large_resources = []
        for i in range(100):
            large_resources.append({
                "service": f"Service{i % 5}",
                "type": "Resource",
                "id": f"resource-{i:03d}",
                "name": f"Resource {i}",
                "region": "us-east-1",
                "compliance_status": "compliant" if i % 2 == 0 else "non_compliant"
            })
            
        large_bom_data = BOMData(
            resources=large_resources,
            network_analysis={"total_vpcs": 10, "total_subnets": 50},
            security_analysis={"total_security_groups": 25, "high_risk_rules": 5},
            compliance_summary={
                "total_resources": 100,
                "compliant_resources": 50,
                "non_compliant_resources": 50,
                "compliance_percentage": 50.0
            },
            generation_metadata={
                "generated_at": datetime.now(timezone.utc).isoformat(),
                "total_resources": 100
            }
        )
        
        builder = ExcelWorkbookBuilder(self.test_config)
        output_path = os.path.join(self.temp_dir, "large_test.xlsx")
        
        result = builder.generate_document(large_bom_data, output_path)
        
        self.assertTrue(result.success)
        self.assertTrue(os.path.exists(output_path))
        self.assertGreater(os.path.getsize(output_path), 10000)  # Should be reasonably large file


if __name__ == '__main__':
    unittest.main()