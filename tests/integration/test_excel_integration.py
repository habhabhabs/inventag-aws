#!/usr/bin/env python3
"""
Integration tests for Excel document generation.

Tests the complete Excel document generation workflow.
"""

import unittest
import tempfile
import os
import shutil
from datetime import datetime, timezone

from inventag.reporting.document_generator import DocumentGenerator, DocumentConfig, BrandingConfig
from inventag.reporting.bom_processor import BOMData

# Check if openpyxl is available
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class TestExcelIntegration(unittest.TestCase):
    """Integration test cases for Excel document generation."""
    
    def setUp(self):
        """Set up test fixtures."""
        self.temp_dir = tempfile.mkdtemp()
        
        # Create comprehensive test BOM data
        self.test_bom_data = BOMData(
            resources=[
                {
                    "service": "EC2",
                    "type": "Instance",
                    "id": "i-1234567890abcdef0",
                    "name": "web-server-1",
                    "region": "us-east-1",
                    "account_id": "123456789012",
                    "arn": "arn:aws:ec2:us-east-1:123456789012:instance/i-1234567890abcdef0",
                    "compliance_status": "compliant",
                    "tags": {
                        "Name": "web-server-1",
                        "Environment": "production",
                        "inventag:remarks": "Primary web server",
                        "inventag:costcenter": "IT-001"
                    },
                    "vpc_id": "vpc-12345678",
                    "subnet_id": "subnet-12345678",
                    "security_groups": ["sg-12345678"],
                    "instance_type": "t3.medium",
                    "state": "running"
                },
                {
                    "service": "S3",
                    "type": "Bucket",
                    "id": "test-bucket-12345",
                    "name": "test-bucket-12345",
                    "region": "us-east-1",
                    "account_id": "123456789012",
                    "arn": "arn:aws:s3:::test-bucket-12345",
                    "compliance_status": "non_compliant",
                    "tags": {
                        "Name": "test-bucket-12345",
                        "Environment": "development"
                    },
                    "encryption": "AES256",
                    "versioning": "Enabled",
                    "public_access_blocked": True
                },
                {
                    "service": "RDS",
                    "type": "DBInstance",
                    "id": "database-1",
                    "name": "database-1",
                    "region": "us-west-2",
                    "account_id": "123456789012",
                    "arn": "arn:aws:rds:us-west-2:123456789012:db:database-1",
                    "compliance_status": "compliant",
                    "tags": {
                        "Name": "database-1",
                        "Environment": "production",
                        "inventag:remarks": "Main application database",
                        "inventag:costcenter": "IT-001"
                    },
                    "engine": "mysql",
                    "engine_version": "8.0.35",
                    "instance_class": "db.t3.micro",
                    "allocated_storage": 20
                },
                {
                    "service": "LAMBDA",
                    "type": "Function",
                    "id": "data-processor",
                    "name": "data-processor",
                    "region": "us-east-1",
                    "account_id": "123456789012",
                    "arn": "arn:aws:lambda:us-east-1:123456789012:function:data-processor",
                    "compliance_status": "compliant",
                    "tags": {
                        "Name": "data-processor",
                        "Environment": "production",
                        "inventag:costcenter": "DEV-001"
                    },
                    "runtime": "python3.9",
                    "memory_size": 256,
                    "timeout": 30
                },
                {
                    "service": "VPC",
                    "type": "SecurityGroup",
                    "id": "sg-12345678",
                    "name": "web-server-sg",
                    "region": "us-east-1",
                    "account_id": "123456789012",
                    "arn": "arn:aws:ec2:us-east-1:123456789012:security-group/sg-12345678",
                    "compliance_status": "non_compliant",
                    "tags": {
                        "Name": "web-server-sg",
                        "Environment": "production"
                    },
                    "vpc_id": "vpc-12345678",
                    "group_name": "web-server-sg",
                    "description": "Security group for web servers"
                }
            ],
            network_analysis={
                "total_vpcs": 2,
                "total_subnets": 6,
                "vpc_utilization": {
                    "vpc-12345678": {
                        "name": "main-vpc",
                        "cidr_block": "10.0.0.0/16",
                        "utilization_percentage": 25.5,
                        "available_ips": 65280
                    },
                    "vpc-87654321": {
                        "name": "dev-vpc",
                        "cidr_block": "10.1.0.0/16",
                        "utilization_percentage": 15.2,
                        "available_ips": 65400
                    }
                }
            },
            security_analysis={
                "total_security_groups": 8,
                "high_risk_rules": 2,
                "overly_permissive_rules": [
                    {
                        "group_id": "sg-12345678",
                        "rule": "0.0.0.0/0:22",
                        "risk_level": "high"
                    },
                    {
                        "group_id": "sg-87654321",
                        "rule": "0.0.0.0/0:3389",
                        "risk_level": "high"
                    }
                ]
            },
            compliance_summary={
                "total_resources": 5,
                "compliant_resources": 3,
                "non_compliant_resources": 2,
                "compliance_percentage": 60.0
            },
            generation_metadata={
                "generated_at": datetime.now(timezone.utc).isoformat(),
                "total_resources": 5,
                "processing_time_seconds": 3.2,
                "data_source": "aws_resource_inventory.py"
            },
            custom_attributes=["inventag:remarks", "inventag:costcenter"]
        )
        
    def tearDown(self):
        """Clean up test fixtures."""
        shutil.rmtree(self.temp_dir, ignore_errors=True)
        
    @unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl not available")
    def test_excel_document_generation_integration(self):
        """Test complete Excel document generation workflow."""
        config = DocumentConfig(
            output_formats=["excel"],
            output_directory=self.temp_dir,
            filename_template="integration_test_{timestamp}",
            branding=BrandingConfig(
                company_name="Integration Test Corp",
                color_scheme={
                    "primary": "2E75B6",
                    "accent": "70AD47",
                    "danger": "C5504B",
                    "warning": "FFC000"
                }
            ),
            validate_before_generation=True
        )
        
        generator = DocumentGenerator(config)
        
        # Generate documents
        summary = generator.generate_bom_documents(self.test_bom_data)
        
        # Verify generation summary
        self.assertEqual(summary.total_formats, 1)
        self.assertEqual(summary.successful_formats, 1)
        self.assertEqual(summary.failed_formats, 0)
        self.assertEqual(len(summary.results), 1)
        
        result = summary.results[0]
        self.assertTrue(result.success)
        self.assertEqual(result.format_type, "excel")
        self.assertTrue(result.filename.endswith(".xlsx"))
        self.assertIsNone(result.error_message)
        self.assertGreater(result.generation_time_seconds, 0)
        
        # Verify file was created
        output_file = os.path.join(self.temp_dir, result.filename)
        self.assertTrue(os.path.exists(output_file))
        self.assertGreater(os.path.getsize(output_file), 10000)  # Should be substantial file
        
        # Load and verify Excel content
        wb = openpyxl.load_workbook(output_file)
        
        # Verify expected sheets exist
        expected_sheets = [
            "Executive Summary",
            "EC2",
            "S3", 
            "RDS",
            "LAMBDA",
            "VPC",
            "Network Analysis",
            "Security Analysis",
            "Compliance Details"
        ]
        
        for sheet_name in expected_sheets:
            self.assertIn(sheet_name, wb.sheetnames, f"Sheet '{sheet_name}' not found")
            
        # Verify Executive Summary content
        exec_summary = wb["Executive Summary"]
        self.assertEqual(exec_summary["A1"].value, "Executive Summary Dashboard")
        
        # Verify service sheets have data
        ec2_sheet = wb["EC2"]
        self.assertEqual(ec2_sheet["A1"].value, "EC2 Resources")
        
        s3_sheet = wb["S3"]
        self.assertEqual(s3_sheet["A1"].value, "S3 Resources")
        
        # Verify Network Analysis sheet
        network_sheet = wb["Network Analysis"]
        self.assertEqual(network_sheet["A1"].value, "Network Analysis & Capacity Planning")
        
        # Verify Security Analysis sheet
        security_sheet = wb["Security Analysis"]
        self.assertEqual(security_sheet["A1"].value, "Security Analysis & Risk Assessment")
        
        # Verify Compliance Details sheet
        compliance_sheet = wb["Compliance Details"]
        self.assertEqual(compliance_sheet["A1"].value, "Detailed Compliance Information")
        
    @unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl not available")
    def test_excel_with_custom_branding(self):
        """Test Excel generation with custom branding."""
        custom_branding = BrandingConfig(
            company_name="Custom Branding Test",
            color_scheme={
                "primary": "FF6B35",
                "accent": "004E89",
                "danger": "D62828",
                "warning": "F77F00"
            },
            font_family="Arial",
            font_size=10
        )
        
        config = DocumentConfig(
            output_formats=["excel"],
            output_directory=self.temp_dir,
            branding=custom_branding,
            filename_template="branded_test_{timestamp}"
        )
        
        generator = DocumentGenerator(config)
        summary = generator.generate_bom_documents(self.test_bom_data)
        
        # Should succeed
        self.assertEqual(summary.successful_formats, 1)
        result = summary.results[0]
        self.assertTrue(result.success)
        
        # Verify file and branding
        output_file = os.path.join(self.temp_dir, result.filename)
        wb = openpyxl.load_workbook(output_file)
        
        # Check workbook properties
        self.assertEqual(wb.properties.title, "Custom Branding Test - Cloud BOM Report")
        self.assertEqual(wb.properties.creator, "InvenTag Cloud BOM Generator")
        
    @unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl not available")
    def test_excel_with_large_dataset(self):
        """Test Excel generation with larger dataset."""
        # Create larger test dataset
        large_resources = []
        services = ["EC2", "S3", "RDS", "LAMBDA", "VPC", "ELB", "CLOUDFORMATION"]
        
        for i in range(50):
            service = services[i % len(services)]
            large_resources.append({
                "service": service,
                "type": f"Resource{i % 3}",
                "id": f"resource-{i:03d}",
                "name": f"Resource {i}",
                "region": "us-east-1" if i % 2 == 0 else "us-west-2",
                "account_id": "123456789012",
                "compliance_status": "compliant" if i % 3 == 0 else "non_compliant",
                "tags": {
                    "Name": f"Resource {i}",
                    "Environment": "production" if i % 2 == 0 else "development",
                    "inventag:costcenter": f"CC-{i % 5:03d}"
                }
            })
            
        large_bom_data = BOMData(
            resources=large_resources,
            network_analysis={
                "total_vpcs": 5,
                "total_subnets": 20,
                "vpc_utilization": {
                    f"vpc-{i:08d}": {
                        "name": f"vpc-{i}",
                        "cidr_block": f"10.{i}.0.0/16",
                        "utilization_percentage": (i * 10) % 100,
                        "available_ips": 65536 - (i * 100)
                    }
                    for i in range(5)
                }
            },
            security_analysis={
                "total_security_groups": 15,
                "high_risk_rules": 3,
                "overly_permissive_rules": [
                    {
                        "group_id": f"sg-{i:08d}",
                        "rule": "0.0.0.0/0:22",
                        "risk_level": "high"
                    }
                    for i in range(3)
                ]
            },
            compliance_summary={
                "total_resources": 50,
                "compliant_resources": 17,  # Every 3rd resource
                "non_compliant_resources": 33,
                "compliance_percentage": 34.0
            },
            generation_metadata={
                "generated_at": datetime.now(timezone.utc).isoformat(),
                "total_resources": 50,
                "processing_time_seconds": 5.8
            }
        )
        
        config = DocumentConfig(
            output_formats=["excel"],
            output_directory=self.temp_dir,
            filename_template="large_dataset_test_{timestamp}"
        )
        
        generator = DocumentGenerator(config)
        summary = generator.generate_bom_documents(large_bom_data)
        
        # Should succeed
        self.assertEqual(summary.successful_formats, 1)
        result = summary.results[0]
        self.assertTrue(result.success)
        
        # Verify file size is reasonable for large dataset
        output_file = os.path.join(self.temp_dir, result.filename)
        self.assertTrue(os.path.exists(output_file))
        self.assertGreater(os.path.getsize(output_file), 20000)  # Should be larger file
        
        # Verify workbook structure
        wb = openpyxl.load_workbook(output_file)
        
        # Should have sheets for all services
        for service in services:
            self.assertIn(service, wb.sheetnames)
            
    @unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl not available")
    def test_excel_error_recovery(self):
        """Test Excel generation error recovery."""
        # Create invalid BOM data to trigger validation error
        invalid_bom_data = BOMData(
            resources=[],  # Empty resources should trigger validation error
            generation_metadata={}
        )
        
        config = DocumentConfig(
            output_formats=["excel"],
            output_directory=self.temp_dir,
            filename_template="error_test_{timestamp}",
            validate_before_generation=True
        )
        
        generator = DocumentGenerator(config)
        
        # Should raise DocumentGenerationError due to validation failure
        with self.assertRaises(Exception):  # DocumentGenerationError
            generator.generate_bom_documents(invalid_bom_data)
        
    @unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl not available")
    def test_excel_with_minimal_data(self):
        """Test Excel generation with minimal data."""
        minimal_bom_data = BOMData(
            resources=[
                {
                    "service": "EC2",
                    "type": "Instance",
                    "id": "i-minimal",
                    "compliance_status": "compliant"
                }
            ],
            network_analysis={},
            security_analysis={},
            compliance_summary={
                "total_resources": 1,
                "compliant_resources": 1,
                "non_compliant_resources": 0,
                "compliance_percentage": 100.0
            },
            generation_metadata={
                "generated_at": datetime.now(timezone.utc).isoformat(),
                "total_resources": 1
            }
        )
        
        config = DocumentConfig(
            output_formats=["excel"],
            output_directory=self.temp_dir,
            filename_template="minimal_test_{timestamp}"
        )
        
        generator = DocumentGenerator(config)
        summary = generator.generate_bom_documents(minimal_bom_data)
        
        # Should succeed even with minimal data
        self.assertEqual(summary.successful_formats, 1)
        result = summary.results[0]
        self.assertTrue(result.success)
        
        # Verify file was created
        output_file = os.path.join(self.temp_dir, result.filename)
        self.assertTrue(os.path.exists(output_file))
        
        # Verify basic structure
        wb = openpyxl.load_workbook(output_file)
        self.assertIn("Executive Summary", wb.sheetnames)
        self.assertIn("EC2", wb.sheetnames)


if __name__ == '__main__':
    unittest.main()